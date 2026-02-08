"""Excel Action Plan generator for Google Sheets compatibility."""

from io import BytesIO

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)

from utils.logger import get_logger

log = get_logger("excel")

# ─── Style Constants ────────────────────────────────────────────

BLUE = "2563EB"
BLUE_LIGHT = "EFF6FF"
WHITE = "FFFFFF"
GRAY_BG = "F9FAFB"
GRAY_BORDER = "E5E7EB"
GREEN_BG = "ECFDF5"
GREEN_TEXT = "059669"
RED_TEXT = "DC2626"
ORANGE_TEXT = "EA580C"
YELLOW_TEXT = "CA8A04"

HEADER_FILL = PatternFill("solid", fgColor=BLUE)
HEADER_FONT = Font(name="Inter", bold=True, color=WHITE, size=11)
BODY_FONT = Font(name="Inter", size=10)
BOLD_FONT = Font(name="Inter", size=10, bold=True)
LINK_FONT = Font(name="Inter", size=10, color=BLUE)
THIN_BORDER = Border(bottom=Side(style="thin", color=GRAY_BORDER))

SEVERITY_STYLES = {
    "critical": {
        "fill": PatternFill("solid", fgColor="FEF2F2"),
        "font": Font(name="Inter", size=10, color=RED_TEXT, bold=True),
    },
    "high": {
        "fill": PatternFill("solid", fgColor="FFF7ED"),
        "font": Font(name="Inter", size=10, color=ORANGE_TEXT, bold=True),
    },
    "medium": {
        "fill": PatternFill("solid", fgColor="FEFCE8"),
        "font": Font(name="Inter", size=10, color=YELLOW_TEXT, bold=True),
    },
    "low": {
        "fill": PatternFill("solid", fgColor="F0FDF4"),
        "font": Font(name="Inter", size=10, color=GREEN_TEXT, bold=True),
    },
}

CHECKED_FILL = PatternFill("solid", fgColor=GREEN_BG)
ALT_ROW_FILL = PatternFill("solid", fgColor=GRAY_BG)


# ─── Helpers ────────────────────────────────────────────────────


def _set_header_row(ws, headers: list[tuple[str, int]], row: int = 1) -> None:
    """Write a styled header row."""
    for col, (title, width) in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[row].height = 30


def _apply_row_border(ws, row: int, max_col: int) -> None:
    """Apply bottom border to a row."""
    for col in range(1, max_col + 1):
        ws.cell(row=row, column=col).border = THIN_BORDER


# ─── Main Sheet: Action Plan ───────────────────────────────────


def _write_action_plan(ws, top_5: list, all_findings: dict) -> None:
    """Write the main Action Plan sheet with Top 5 + all findings."""
    headers = [
        ("Priority", 10),
        ("Category", 14),
        ("Issue", 45),
        ("URLs Affected", 14),
        ("What To Do", 55),
        ("Impact", 10),
        ("Effort", 10),
        ("Status", 10),
    ]
    _set_header_row(ws, headers)

    row = 2

    # ── Top 5 Quick Wins ──
    for i, win in enumerate(top_5, 1):
        impact = (win.get("impact") or "medium").lower()
        effort = (win.get("effort") or "medium").lower()

        ws.cell(row=row, column=1, value=f"#{i}").font = BOLD_FONT
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")

        cat_cell = ws.cell(row=row, column=2, value=win.get("category", "").title())
        cat_cell.font = BODY_FONT

        issue_cell = ws.cell(row=row, column=3, value=win.get("issue", ""))
        issue_cell.font = BOLD_FONT
        issue_cell.alignment = Alignment(wrap_text=True)

        urls_cell = ws.cell(row=row, column=4, value=win.get("urls_affected", 0))
        urls_cell.font = BODY_FONT
        urls_cell.alignment = Alignment(horizontal="center")

        action_cell = ws.cell(row=row, column=5, value=win.get("what_to_do", ""))
        action_cell.font = BODY_FONT
        action_cell.alignment = Alignment(wrap_text=True)

        impact_cell = ws.cell(row=row, column=6, value=impact.upper())
        sev_key = {"high": "critical", "medium": "medium", "low": "low"}.get(impact, "medium")
        impact_cell.font = SEVERITY_STYLES.get(sev_key, {}).get("font", BODY_FONT)
        impact_cell.alignment = Alignment(horizontal="center")

        effort_cell = ws.cell(row=row, column=7, value=effort.upper())
        effort_cell.font = BODY_FONT
        effort_cell.alignment = Alignment(horizontal="center")

        # Checkbox using Unicode
        status_cell = ws.cell(row=row, column=8, value="\u2610")  # ☐
        status_cell.font = Font(name="Inter", size=14)
        status_cell.alignment = Alignment(horizontal="center")

        # Highlight row with light blue for top 5
        for col in range(1, 9):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=BLUE_LIGHT)

        _apply_row_border(ws, row, 8)
        ws.row_dimensions[row].height = 45
        row += 1

    # ── Separator row ──
    sep_cell = ws.cell(row=row, column=1, value="ALL FINDINGS")
    sep_cell.font = Font(name="Inter", size=10, bold=True, color="6B7280")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    sep_cell.fill = PatternFill("solid", fgColor="F3F4F6")
    sep_cell.alignment = Alignment(horizontal="center")
    row += 1

    # ── All Findings ──
    for category_name, findings in all_findings.items():
        if not findings:
            continue
        for finding in findings:
            severity = (finding.get("severity") or "medium").lower()

            ws.cell(row=row, column=1, value="").font = BODY_FONT

            ws.cell(row=row, column=2, value=category_name.title()).font = BODY_FONT

            issue_cell = ws.cell(row=row, column=3, value=finding.get("issue", ""))
            issue_cell.font = BODY_FONT
            issue_cell.alignment = Alignment(wrap_text=True)

            ws.cell(row=row, column=4, value=finding.get("count", 0)).font = BODY_FONT
            ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")

            ws.cell(row=row, column=5, value="").font = BODY_FONT

            sev_cell = ws.cell(row=row, column=6, value=severity.upper())
            style = SEVERITY_STYLES.get(severity, {})
            sev_cell.font = style.get("font", BODY_FONT)
            sev_cell.fill = style.get("fill", PatternFill())
            sev_cell.alignment = Alignment(horizontal="center")

            ws.cell(row=row, column=7, value="").font = BODY_FONT

            status_cell = ws.cell(row=row, column=8, value="\u2610")
            status_cell.font = Font(name="Inter", size=14)
            status_cell.alignment = Alignment(horizontal="center")

            # Alternating row colors
            if row % 2 == 0:
                for col in range(1, 9):
                    if not ws.cell(row=row, column=col).fill or ws.cell(row=row, column=col).fill.fgColor.rgb == "00000000":
                        ws.cell(row=row, column=col).fill = ALT_ROW_FILL

            _apply_row_border(ws, row, 8)
            row += 1

    # Auto filter and freeze
    ws.auto_filter.ref = f"A1:H{row - 1}"
    ws.freeze_panes = "A2"


# ─── Detail Sheets ──────────────────────────────────────────────


def _write_detail_sheet(wb, category: str, findings: list) -> None:
    """Write a detail sheet for a category with all affected URLs."""
    sheet_name = f"{category.title()} Details"[:31]
    ws = wb.create_sheet(sheet_name)

    headers = [
        ("Issue", 35),
        ("Severity", 12),
        ("Affected URL", 60),
        ("Details", 40),
    ]
    _set_header_row(ws, headers)

    row = 2
    for finding in findings:
        issue_name = finding.get("issue", "")
        severity = (finding.get("severity") or "medium").lower()
        urls = finding.get("urls", [])

        for url in urls[:50]:
            ws.cell(row=row, column=1, value=issue_name).font = BODY_FONT
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)

            sev_cell = ws.cell(row=row, column=2, value=severity.upper())
            style = SEVERITY_STYLES.get(severity, {})
            sev_cell.font = style.get("font", BODY_FONT)
            sev_cell.alignment = Alignment(horizontal="center")

            url_cell = ws.cell(row=row, column=3, value=url)
            url_cell.font = LINK_FONT
            url_cell.alignment = Alignment(wrap_text=True)

            ws.cell(row=row, column=4, value="").font = BODY_FONT

            _apply_row_border(ws, row, 4)
            row += 1

    ws.auto_filter.ref = f"A1:D{max(row - 1, 1)}"
    ws.freeze_panes = "A2"


# ─── Public API ─────────────────────────────────────────────────


def create_action_plan(
    top_5: list,
    all_findings: dict,
    domain: str = "",
) -> BytesIO:
    """
    Create the Excel Action Plan workbook.

    Args:
        top_5: List of top 5 quick wins from LLM
        all_findings: Dict with content/headings/links/technical findings
        domain: Domain name for sheet title

    Returns:
        BytesIO with the Excel file
    """
    log.info("Generating Excel action plan...")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Action Plan"

    _write_action_plan(ws, top_5, all_findings)

    # Detail sheets per category
    for category, findings in all_findings.items():
        if findings:
            _write_detail_sheet(wb, category, findings)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    log.info("Excel action plan generated")
    return output
