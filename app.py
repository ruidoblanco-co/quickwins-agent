import streamlit as st
import time
import json
import re
import requests
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse, urljoin
from collections import defaultdict
from bs4 import BeautifulSoup
from io import BytesIO
import google.generativeai as genai
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ───────────────────────────────────────────
# CONFIG
# ───────────────────────────────────────────
st.set_page_config(
    page_title="Quick Wins · SEO Agent",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="collapsed",
)

BASE_DIR = Path(__file__).parent
PROMPTS_DIR = BASE_DIR / "prompts"
PROMPT_QUICKWINS = PROMPTS_DIR / "quickwins.md"
PROMPT_FIX = PROMPTS_DIR / "generate_fix.md"

CRAWL_TIMEOUT = 12
MAX_PAGES = 80
MAX_INTERNAL_LINKS_PER_PAGE = 10
MAX_BROKEN_LINK_CHECKS = 120

# ───────────────────────────────────────────
# API SETUP
# ───────────────────────────────────────────
try:
    GEMINI_API_KEY = st.secrets["GOOGLE_API_KEY"]
    genai.configure(api_key=GEMINI_API_KEY)
    GEMINI_AVAILABLE = True
except Exception:
    GEMINI_AVAILABLE = False

# ───────────────────────────────────────────
# DESIGN SYSTEM (White, clean, Sandy-inspired)
# ───────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:ital,opsz,wght@0,9..40,300;0,9..40,400;0,9..40,500;0,9..40,600;0,9..40,700;1,9..40,400&family=JetBrains+Mono:wght@400;500&display=swap');

:root {
    --bg: #FAFAFA;
    --surface: #FFFFFF;
    --border: #E8E8E8;
    --border-hover: #D0D0D0;
    --text-primary: #1A1A1A;
    --text-secondary: #6B6B6B;
    --text-muted: #9B9B9B;
    --accent: #FF6B35;
    --accent-light: #FFF3ED;
    --accent-hover: #E55A2B;
    --critical: #DC2626;
    --critical-bg: #FEF2F2;
    --high: #EA580C;
    --high-bg: #FFF7ED;
    --medium: #CA8A04;
    --medium-bg: #FEFCE8;
    --low: #16A34A;
    --low-bg: #F0FDF4;
    --mono: 'JetBrains Mono', monospace;
    --sans: 'DM Sans', -apple-system, BlinkMacSystemFont, sans-serif;
}

/* Global resets */
.stApp {
    background-color: var(--bg) !important;
    font-family: var(--sans) !important;
}

[data-testid="stSidebar"] {
    background: var(--surface);
    border-right: 1px solid var(--border);
}

/* Hide Streamlit branding */
#MainMenu, footer, header {visibility: hidden;}
.stDeployButton {display: none;}

/* Typography */
h1, h2, h3, h4, h5, h6 {
    font-family: var(--sans) !important;
    color: var(--text-primary) !important;
    font-weight: 600 !important;
}

p, li, span, div {
    font-family: var(--sans);
}

/* ─── HEADER ─── */
.qw-header {
    text-align: center;
    padding: 3rem 1rem 2rem;
    max-width: 640px;
    margin: 0 auto;
}

.qw-logo {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 56px;
    height: 56px;
    border-radius: 14px;
    background: linear-gradient(135deg, #FF6B35, #FF8F5E);
    color: white;
    font-size: 26px;
    margin-bottom: 1rem;
    box-shadow: 0 4px 16px rgba(255, 107, 53, 0.25);
}

.qw-title {
    font-size: 28px;
    font-weight: 700;
    color: var(--text-primary);
    margin: 0 0 0.35rem;
    letter-spacing: -0.5px;
}

.qw-subtitle {
    font-size: 15px;
    color: var(--text-secondary);
    margin: 0;
    line-height: 1.5;
}

/* ─── INPUT AREA ─── */
.stTextInput > div > div > input {
    background: var(--surface) !important;
    border: 1.5px solid var(--border) !important;
    border-radius: 10px !important;
    padding: 12px 16px !important;
    font-size: 15px !important;
    font-family: var(--sans) !important;
    color: var(--text-primary) !important;
    transition: border-color 0.2s ease;
}

.stTextInput > div > div > input:focus {
    border-color: var(--accent) !important;
    box-shadow: 0 0 0 3px rgba(255, 107, 53, 0.1) !important;
}

.stTextInput > div > div > input::placeholder {
    color: var(--text-muted) !important;
}

/* ─── BUTTONS ─── */
.stButton > button {
    background: var(--accent) !important;
    color: white !important;
    border: none !important;
    border-radius: 10px !important;
    padding: 0.6rem 1.5rem !important;
    font-size: 14px !important;
    font-weight: 600 !important;
    font-family: var(--sans) !important;
    transition: all 0.2s ease !important;
    letter-spacing: 0.01em;
}

.stButton > button:hover {
    background: var(--accent-hover) !important;
    transform: translateY(-1px);
    box-shadow: 0 4px 12px rgba(255, 107, 53, 0.3) !important;
}

.stButton > button:disabled {
    background: var(--border) !important;
    color: var(--text-muted) !important;
    transform: none !important;
    box-shadow: none !important;
}

/* ─── CARDS ─── */
.qw-card {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 12px;
    padding: 1.25rem 1.5rem;
    margin-bottom: 0.75rem;
    transition: border-color 0.2s ease, box-shadow 0.2s ease;
}

.qw-card:hover {
    border-color: var(--border-hover);
    box-shadow: 0 2px 8px rgba(0,0,0,0.04);
}

.qw-card-header {
    display: flex;
    align-items: center;
    gap: 0.75rem;
    margin-bottom: 0.5rem;
}

.qw-card-title {
    font-size: 15px;
    font-weight: 600;
    color: var(--text-primary);
    margin: 0;
    flex: 1;
}

.qw-card-desc {
    font-size: 13.5px;
    color: var(--text-secondary);
    line-height: 1.55;
    margin: 0 0 0.75rem;
}

.qw-card-meta {
    display: flex;
    align-items: center;
    gap: 0.75rem;
    flex-wrap: wrap;
}

/* ─── BADGES ─── */
.badge {
    display: inline-flex;
    align-items: center;
    padding: 3px 10px;
    border-radius: 6px;
    font-size: 11.5px;
    font-weight: 600;
    letter-spacing: 0.03em;
    text-transform: uppercase;
    font-family: var(--mono);
}

.badge-critical { background: var(--critical-bg); color: var(--critical); }
.badge-high { background: var(--high-bg); color: var(--high); }
.badge-medium { background: var(--medium-bg); color: var(--medium); }
.badge-low { background: var(--low-bg); color: var(--low); }

.badge-cat {
    background: #F3F4F6;
    color: var(--text-secondary);
    font-weight: 500;
}

.badge-effort {
    background: #EDE9FE;
    color: #7C3AED;
    font-weight: 500;
}

.badge-urls {
    background: #ECFDF5;
    color: #059669;
    font-weight: 500;
}

/* ─── SNAPSHOT STATS ─── */
.qw-stats {
    display: flex;
    gap: 1rem;
    margin: 1.5rem 0;
    flex-wrap: wrap;
}

.qw-stat {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 10px;
    padding: 1rem 1.25rem;
    flex: 1;
    min-width: 120px;
    text-align: center;
}

.qw-stat-value {
    font-size: 28px;
    font-weight: 700;
    color: var(--text-primary);
    font-family: var(--mono);
    line-height: 1;
}

.qw-stat-label {
    font-size: 12px;
    color: var(--text-muted);
    margin-top: 4px;
    text-transform: uppercase;
    letter-spacing: 0.05em;
    font-weight: 500;
}

.qw-stat-critical .qw-stat-value { color: var(--critical); }
.qw-stat-high .qw-stat-value { color: var(--high); }
.qw-stat-medium .qw-stat-value { color: var(--medium); }
.qw-stat-low .qw-stat-value { color: var(--low); }

/* ─── FIX RESULTS ─── */
.fix-item {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 10px;
    padding: 1rem 1.25rem;
    margin-bottom: 0.5rem;
}

.fix-url {
    font-family: var(--mono);
    font-size: 12px;
    color: var(--text-muted);
    margin-bottom: 6px;
    word-break: break-all;
}

.fix-current {
    font-size: 13px;
    color: var(--critical);
    text-decoration: line-through;
    margin-bottom: 4px;
}

.fix-suggested {
    font-size: 14px;
    color: var(--low);
    font-weight: 500;
    margin-bottom: 4px;
}

.fix-reason {
    font-size: 12px;
    color: var(--text-muted);
    font-style: italic;
}

/* ─── PROGRESS ─── */
.stProgress > div > div > div > div {
    background: linear-gradient(90deg, var(--accent), #FF8F5E) !important;
    border-radius: 10px;
}

/* ─── EXPANDER ─── */
.streamlit-expanderHeader {
    font-family: var(--sans) !important;
    font-size: 13px !important;
    font-weight: 500 !important;
    color: var(--text-secondary) !important;
    background: transparent !important;
    border: none !important;
}

/* ─── DIVIDER ─── */
.qw-divider {
    height: 1px;
    background: var(--border);
    margin: 2rem 0;
}

/* ─── FOOTER ─── */
.qw-footer {
    text-align: center;
    padding: 2rem 0 1rem;
    font-size: 12px;
    color: var(--text-muted);
}

/* Tabs */
.stTabs [data-baseweb="tab-list"] {
    gap: 0;
    border-bottom: 1px solid var(--border);
}

.stTabs [data-baseweb="tab"] {
    font-family: var(--sans) !important;
    font-size: 13px !important;
    font-weight: 500 !important;
    color: var(--text-secondary) !important;
    padding: 0.5rem 1.25rem !important;
    border-bottom: 2px solid transparent !important;
}

.stTabs [aria-selected="true"] {
    color: var(--accent) !important;
    border-bottom-color: var(--accent) !important;
}

/* Download button */
.stDownloadButton > button {
    background: var(--surface) !important;
    color: var(--text-primary) !important;
    border: 1.5px solid var(--border) !important;
    border-radius: 10px !important;
    font-family: var(--sans) !important;
    font-weight: 500 !important;
    font-size: 13px !important;
}

.stDownloadButton > button:hover {
    border-color: var(--accent) !important;
    color: var(--accent) !important;
    background: var(--accent-light) !important;
    box-shadow: none !important;
    transform: none !important;
}

/* Alert boxes */
.stAlert {
    border-radius: 10px !important;
    font-family: var(--sans) !important;
}
</style>
""", unsafe_allow_html=True)


# ───────────────────────────────────────────
# UTILITIES
# ───────────────────────────────────────────
def normalize_domain(url_or_domain: str) -> str:
    s = (url_or_domain or "").strip()
    if not s:
        return ""
    if s.startswith(("http://", "https://")):
        s = urlparse(s).netloc
    s = s.lower()
    if s.startswith("www."):
        s = s[4:]
    return s.split(":")[0]


def load_prompt(path: Path) -> str:
    return path.read_text(encoding="utf-8") if path.exists() else ""


def strip_json_fences(text: str) -> str:
    t = (text or "").strip()
    t = re.sub(r"^```(?:json)?\s*", "", t, flags=re.IGNORECASE)
    t = re.sub(r"\s*```$", "", t)
    return t.strip()


def safe_int(x, default=0):
    try:
        return int(x)
    except Exception:
        return default


# ───────────────────────────────────────────
# CRAWLER (adapted from Claudio, optimized for quick wins)
# ───────────────────────────────────────────
HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}


def fetch_url(url, timeout=CRAWL_TIMEOUT):
    try:
        return requests.get(url, headers=HEADERS, timeout=timeout, allow_redirects=True)
    except Exception:
        return None


def get_robots_sitemaps(base_url):
    r = fetch_url(urljoin(base_url.rstrip("/") + "/", "robots.txt"))
    if not r or r.status_code >= 400:
        return []
    sitemaps = []
    for line in r.text.splitlines():
        if line.lower().startswith("sitemap:"):
            sm = line.split(":", 1)[1].strip()
            if sm:
                sitemaps.append(sm)
    return list(dict.fromkeys(sitemaps))


def try_default_sitemaps(base_url):
    base = base_url.rstrip("/") + "/"
    return [urljoin(base, "sitemap.xml"), urljoin(base, "sitemap_index.xml")]


def parse_sitemap_xml(xml_text):
    urls, sitemaps = [], []
    try:
        root = ET.fromstring(xml_text)
    except Exception:
        return urls, sitemaps
    tag_end = lambda el, name: el.tag.lower().endswith(name)
    if tag_end(root, "sitemapindex"):
        for el in root.findall(".//"):
            if tag_end(el, "loc") and el.text:
                sitemaps.append(el.text.strip())
    else:
        for el in root.findall(".//"):
            if tag_end(el, "loc") and el.text:
                urls.append(el.text.strip())
    return urls, sitemaps


def fetch_sitemap_urls(sitemap_url, max_urls=6000):
    r = fetch_url(sitemap_url)
    if not r or r.status_code >= 400:
        return []
    urls, sitemaps = parse_sitemap_xml(r.text)
    all_urls = list(urls)
    for sm in (sitemaps or [])[:20]:
        time.sleep(0.1)
        all_urls.extend(fetch_sitemap_urls(sm, max_urls=max_urls))
        if len(all_urls) >= max_urls:
            break
    return list(dict.fromkeys(all_urls))[:max_urls]


def pick_sample(urls, homepage, max_pages=MAX_PAGES):
    urls = [u for u in urls if isinstance(u, str) and u.startswith(("http://", "https://"))]
    urls = list(dict.fromkeys(urls))
    sample = [homepage] if homepage else []
    buckets = defaultdict(list)
    for u in urls:
        try:
            path = (urlparse(u).path or "/").strip("/")
            bucket = path.split("/")[0] if path else "_root"
            buckets[bucket].append(u)
        except Exception:
            continue
    for _, lst in buckets.items():
        if len(sample) >= max_pages:
            break
        if lst[0] not in sample:
            sample.append(lst[0])
    for u in urls:
        if len(sample) >= max_pages:
            break
        if u not in sample:
            sample.append(u)
    return sample[:max_pages]


def extract_signals(url, base_domain):
    r = fetch_url(url)
    if not r:
        return {"url": url, "status": None, "error": "request_failed"}
    if "text/html" not in (r.headers.get("Content-Type") or "").lower():
        return {"url": url, "status": r.status_code, "error": "non_html"}

    soup = BeautifulSoup(r.text, "html.parser")
    title = soup.title.string.strip() if soup.title and soup.title.string else ""
    meta = ""
    md = soup.find("meta", attrs={"name": "description"})
    if md:
        meta = (md.get("content") or "").strip()

    canonical = ""
    c = soup.find("link", attrs={"rel": lambda x: x and "canonical" in x.lower()})
    if c:
        canonical = (c.get("href") or "").strip()

    robots_meta = ""
    rm = soup.find("meta", attrs={"name": lambda x: x and x.lower() == "robots"})
    if rm:
        robots_meta = (rm.get("content") or "").strip().lower()

    imgs = soup.find_all("img")
    internal_links = []
    for a in soup.find_all("a", href=True):
        href = (a.get("href") or "").strip()
        if not href or href.startswith(("#", "mailto:", "tel:", "javascript:")):
            continue
        abs_url = href if href.startswith(("http://", "https://")) else urljoin(r.url, href)
        if normalize_domain(abs_url) == base_domain:
            internal_links.append(abs_url)
        if len(internal_links) >= MAX_INTERNAL_LINKS_PER_PAGE:
            break

    return {
        "url": url,
        "final_url": r.url,
        "status": r.status_code,
        "title": title,
        "title_len": len(title),
        "meta": meta,
        "meta_len": len(meta),
        "canonical": canonical,
        "robots_meta": robots_meta,
        "h1_count": len(soup.find_all("h1")),
        "word_count": len(soup.get_text(" ", strip=True).split()),
        "images_total": len(imgs),
        "images_missing_alt": sum(1 for img in imgs if not (img.get("alt") or "").strip()),
        "hreflang_count": len(soup.find_all("link", attrs={"rel": lambda x: x and "alternate" in x.lower(), "hreflang": True})),
        "jsonld_count": len(soup.find_all("script", attrs={"type": "application/ld+json"})),
        "sample_internal_links": internal_links,
    }


def check_broken_links(links):
    broken = []
    for link in links:
        try:
            r = requests.head(link, headers=HEADERS, timeout=CRAWL_TIMEOUT, allow_redirects=True)
            code = r.status_code
            if code >= 400:
                r2 = requests.get(link, headers=HEADERS, timeout=CRAWL_TIMEOUT, allow_redirects=True)
                code = r2.status_code
            if code >= 400:
                broken.append({"url": link, "status": code})
        except Exception:
            broken.append({"url": link, "status": None})
        time.sleep(0.04)
        if len(broken) >= 20:
            break
    return broken


def build_findings(pages, base_domain):
    summary = {
        "analyzed_pages": len(pages),
        "status_4xx_5xx": 0, "redirects": 0,
        "missing_title": 0, "missing_meta": 0,
        "missing_h1": 0, "multiple_h1": 0,
        "noindex_pages": 0, "missing_canonical": 0,
        "canonical_mismatch": 0, "thin_pages_lt_250w": 0,
        "total_images_missing_alt": 0,
        "pages_with_schema": 0, "pages_with_hreflang": 0,
    }
    examples = {
        "duplicate_titles": [], "duplicate_meta": [],
        "noindex_examples": [], "canonical_examples": [],
        "thin_examples": [], "status_examples": [],
    }
    titles, metas = defaultdict(list), defaultdict(list)

    for p in pages:
        status = p.get("status")
        url = p.get("final_url") or p.get("url")

        if status is None or (isinstance(status, int) and status >= 400):
            summary["status_4xx_5xx"] += 1
            if len(examples["status_examples"]) < 10:
                examples["status_examples"].append({"url": url, "status": status})
        elif (p.get("final_url") or p.get("url")) != p.get("url"):
            summary["redirects"] += 1

        title = (p.get("title") or "").strip()
        meta = (p.get("meta") or "").strip()

        if not title:
            summary["missing_title"] += 1
        else:
            titles[title].append(url)
        if not meta:
            summary["missing_meta"] += 1
        else:
            metas[meta].append(url)

        h1c = safe_int(p.get("h1_count"), 0)
        if h1c == 0:
            summary["missing_h1"] += 1
        elif h1c > 1:
            summary["multiple_h1"] += 1

        robots = (p.get("robots_meta") or "").lower()
        if "noindex" in robots:
            summary["noindex_pages"] += 1
            if len(examples["noindex_examples"]) < 10:
                examples["noindex_examples"].append({"url": url, "robots": robots})

        canonical = (p.get("canonical") or "").strip()
        if not canonical:
            summary["missing_canonical"] += 1
        else:
            try:
                c_abs = canonical if canonical.startswith("http") else urljoin(url, canonical)
                if normalize_domain(c_abs) != base_domain:
                    summary["canonical_mismatch"] += 1
                    if len(examples["canonical_examples"]) < 10:
                        examples["canonical_examples"].append({"url": url, "canonical": canonical})
            except Exception:
                pass

        wc = safe_int(p.get("word_count"), 0)
        if 0 < wc < 250:
            summary["thin_pages_lt_250w"] += 1
            if len(examples["thin_examples"]) < 10:
                examples["thin_examples"].append({"url": url, "word_count": wc})

        summary["total_images_missing_alt"] += safe_int(p.get("images_missing_alt"), 0)
        if safe_int(p.get("jsonld_count"), 0) > 0:
            summary["pages_with_schema"] += 1
        if safe_int(p.get("hreflang_count"), 0) > 0:
            summary["pages_with_hreflang"] += 1

    for t, urls in sorted(titles.items(), key=lambda x: len(x[1]), reverse=True):
        if len(urls) > 1 and len(examples["duplicate_titles"]) < 5:
            examples["duplicate_titles"].append({"value": t[:140], "count": len(urls), "urls": urls[:5]})
    for m, urls in sorted(metas.items(), key=lambda x: len(x[1]), reverse=True):
        if len(urls) > 1 and len(examples["duplicate_meta"]) < 5:
            examples["duplicate_meta"].append({"value": m[:160], "count": len(urls), "urls": urls[:5]})

    return summary, examples


def run_crawl(url_input, progress_cb=None):
    base_domain = normalize_domain(url_input)
    if not base_domain:
        raise ValueError("Invalid URL/domain")
    base_url = url_input if url_input.startswith(("http://", "https://")) else "https://" + url_input
    p = urlparse(base_url)
    base_url = f"{p.scheme}://{p.netloc}"

    if progress_cb:
        progress_cb(10, "Checking robots.txt & sitemaps...")

    sitemaps = get_robots_sitemaps(base_url) or try_default_sitemaps(base_url)
    discovered = []
    used_sm = None
    for sm in sitemaps:
        urls = fetch_sitemap_urls(sm, max_urls=6000)
        if urls:
            discovered = urls
            used_sm = sm
            break

    homepage = base_url
    if not discovered:
        sample_urls = [homepage]
        method = "homepage_only"
        disc_count = 1
    else:
        sample_urls = pick_sample(discovered, homepage)
        method = f"sitemap ({used_sm})"
        disc_count = len(discovered)

    if progress_cb:
        progress_cb(20, f"Found {disc_count} URLs. Scanning {len(sample_urls)} pages...")

    pages = []
    for i, u in enumerate(sample_urls):
        pages.append(extract_signals(u, base_domain))
        time.sleep(0.1)
        if progress_cb and i % 5 == 0:
            pct = 20 + int(40 * (i / len(sample_urls)))
            progress_cb(pct, f"Scanning page {i+1}/{len(sample_urls)}...")

    if progress_cb:
        progress_cb(65, "Analyzing findings...")

    summary, examples = build_findings(pages, base_domain)

    # Broken link check (sample)
    if progress_cb:
        progress_cb(70, "Checking internal links...")
    all_links = []
    for pg in pages:
        for ln in (pg.get("sample_internal_links") or []):
            if ln not in all_links:
                all_links.append(ln)
            if len(all_links) >= MAX_BROKEN_LINK_CHECKS:
                break
        if len(all_links) >= MAX_BROKEN_LINK_CHECKS:
            break

    broken = check_broken_links(all_links)
    summary["broken_internal_links_checked"] = len(all_links)
    summary["broken_internal_links_found"] = len(broken)
    examples["broken_links"] = broken

    return {
        "domain": base_domain,
        "audit_date": datetime.now().strftime("%Y-%m-%d"),
        "discovery_method": method,
        "urls_discovered": disc_count,
        "urls_analyzed": len(pages),
        "crawl_summary": summary,
        "pages": pages,
        "examples": examples,
    }


# ───────────────────────────────────────────
# LLM
# ───────────────────────────────────────────
def call_gemini(prompt_text):
    model = genai.GenerativeModel("gemini-2.5-flash")
    resp = model.generate_content(prompt_text)
    return (getattr(resp, "text", "") or "").strip()


def analyze_quickwins(context):
    tmpl = load_prompt(PROMPT_QUICKWINS)
    if not tmpl:
        return None
    prompt = tmpl.replace("{{CONTEXT_JSON}}", json.dumps(context, ensure_ascii=False, indent=2))
    raw = call_gemini(prompt)
    raw = strip_json_fences(raw)
    try:
        return json.loads(raw)
    except Exception:
        m = re.search(r"\{.*\}", raw, re.DOTALL)
        if m:
            try:
                return json.loads(m.group())
            except Exception:
                pass
    return None


def generate_fixes(category, affected_pages):
    tmpl = load_prompt(PROMPT_FIX)
    if not tmpl:
        return None
    fix_context = json.dumps({"category": category, "affected_pages": affected_pages}, ensure_ascii=False, indent=2)
    prompt = tmpl.replace("{{FIX_CONTEXT}}", fix_context)
    raw = call_gemini(prompt)
    raw = strip_json_fences(raw)
    try:
        return json.loads(raw)
    except Exception:
        m = re.search(r"\{.*\}", raw, re.DOTALL)
        if m:
            try:
                return json.loads(m.group())
            except Exception:
                pass
    return None


# ───────────────────────────────────────────
# EXCEL EXPORT
# ───────────────────────────────────────────
def create_excel(qw_data, fixes_data=None):
    wb = openpyxl.Workbook()

    # ─── Overview sheet ───
    ws = wb.active
    ws.title = "Overview"

    accent_fill = PatternFill("solid", fgColor="FF6B35")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell_font = Font(name="Arial", size=10)
    border = Border(
        bottom=Side(style="thin", color="E8E8E8"),
    )

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 60
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 60

    headers = ["#", "Category", "Severity", "Issue", "Description", "URLs", "Effort", "Fix Instruction"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = accent_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    severity_fills = {
        "critical": PatternFill("solid", fgColor="FEF2F2"),
        "high": PatternFill("solid", fgColor="FFF7ED"),
        "medium": PatternFill("solid", fgColor="FEFCE8"),
        "low": PatternFill("solid", fgColor="F0FDF4"),
    }
    severity_fonts = {
        "critical": Font(name="Arial", size=10, color="DC2626", bold=True),
        "high": Font(name="Arial", size=10, color="EA580C", bold=True),
        "medium": Font(name="Arial", size=10, color="CA8A04", bold=True),
        "low": Font(name="Arial", size=10, color="16A34A", bold=True),
    }

    for i, qw in enumerate((qw_data.get("quick_wins") or []), start=1):
        row = i + 1
        sev = (qw.get("severity") or "medium").lower()

        ws.cell(row=row, column=1, value=i).font = cell_font
        ws.cell(row=row, column=2, value=qw.get("category", "")).font = cell_font
        sc = ws.cell(row=row, column=3, value=sev.upper())
        sc.font = severity_fonts.get(sev, cell_font)
        sc.fill = severity_fills.get(sev, PatternFill())
        sc.alignment = Alignment(horizontal="center")

        ws.cell(row=row, column=4, value=qw.get("title", "")).font = Font(name="Arial", size=10, bold=True)
        ws.cell(row=row, column=5, value=qw.get("description", "")).font = cell_font
        ws.cell(row=row, column=6, value=qw.get("affected_urls_count", 0)).font = cell_font
        ws.cell(row=row, column=6).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=7, value=qw.get("effort", "M")).font = cell_font
        ws.cell(row=row, column=7).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=8, value=qw.get("fix_instruction", "")).font = cell_font

        for col in range(1, 9):
            ws.cell(row=row, column=col).border = border

    ws.auto_filter.ref = f"A1:H{len(qw_data.get('quick_wins', [])) + 1}"
    ws.freeze_panes = "A2"

    # ─── Fixes sheets (if any) ───
    if fixes_data:
        for cat, fixes in fixes_data.items():
            if not fixes:
                continue
            sheet_name = f"Fixes - {cat.title()}"[:31]
            ws2 = wb.create_sheet(sheet_name)
            ws2.column_dimensions["A"].width = 50
            ws2.column_dimensions["B"].width = 40
            ws2.column_dimensions["C"].width = 50
            ws2.column_dimensions["D"].width = 50

            fix_headers = ["URL", "Current Value", "Suggested Fix", "Reasoning"]
            for col, h in enumerate(fix_headers, 1):
                c = ws2.cell(row=1, column=col, value=h)
                c.fill = accent_fill
                c.font = header_font
                c.alignment = Alignment(horizontal="center", vertical="center")

            for j, fix in enumerate(fixes, start=1):
                ws2.cell(row=j+1, column=1, value=fix.get("url", "")).font = cell_font
                ws2.cell(row=j+1, column=2, value=fix.get("current_value", "")).font = Font(name="Arial", size=10, color="999999")
                ws2.cell(row=j+1, column=3, value=fix.get("suggested_fix", "")).font = Font(name="Arial", size=10, color="16A34A", bold=True)
                ws2.cell(row=j+1, column=4, value=fix.get("reasoning", "")).font = cell_font

            ws2.freeze_panes = "A2"

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ───────────────────────────────────────────
# RENDER HELPERS
# ───────────────────────────────────────────
def severity_badge(sev):
    s = (sev or "medium").lower()
    return f'<span class="badge badge-{s}">{s}</span>'


def render_card(qw, idx):
    sev = (qw.get("severity") or "medium").lower()
    cat = qw.get("category", "general")
    effort = qw.get("effort", "M")
    count = qw.get("affected_urls_count", 0)

    return f"""
    <div class="qw-card">
        <div class="qw-card-header">
            <span class="badge badge-{sev}">{sev}</span>
            <p class="qw-card-title">{qw.get('title', 'Quick Win')}</p>
        </div>
        <p class="qw-card-desc">{qw.get('description', '')}</p>
        <div class="qw-card-meta">
            <span class="badge badge-cat">{cat}</span>
            <span class="badge badge-urls">{count} URLs</span>
            <span class="badge badge-effort">Effort: {effort}</span>
        </div>
    </div>
    """


# ───────────────────────────────────────────
# SESSION STATE
# ───────────────────────────────────────────
if "qw_result" not in st.session_state:
    st.session_state.qw_result = None
if "crawl_context" not in st.session_state:
    st.session_state.crawl_context = None
if "fixes" not in st.session_state:
    st.session_state.fixes = {}


# ───────────────────────────────────────────
# UI
# ───────────────────────────────────────────

# Header
st.markdown("""
<div class="qw-header">
    <div class="qw-logo">⚡</div>
    <h1 class="qw-title">Quick Wins</h1>
    <p class="qw-subtitle">Paste a URL. Get a prioritized list of SEO fixes you can do today.</p>
</div>
""", unsafe_allow_html=True)

# Input
col_input, col_btn = st.columns([4, 1])
with col_input:
    url_input = st.text_input(
        "Website URL",
        placeholder="https://example.com",
        label_visibility="collapsed",
    )
with col_btn:
    run_btn = st.button("Scan →", disabled=not url_input or not GEMINI_AVAILABLE, use_container_width=True)

if not GEMINI_AVAILABLE:
    st.error("Gemini API key not configured. Add `GOOGLE_API_KEY` to Streamlit secrets.")

# ─── Run scan ───
if run_btn and url_input:
    st.session_state.qw_result = None
    st.session_state.crawl_context = None
    st.session_state.fixes = {}

    progress = st.progress(0)
    status = st.empty()

    def update_progress(pct, msg):
        progress.progress(min(pct, 100))
        status.caption(msg)

    try:
        context = run_crawl(url_input, progress_cb=update_progress)
        st.session_state.crawl_context = context

        update_progress(80, "AI is analyzing quick wins...")
        result = analyze_quickwins(context)

        if result:
            st.session_state.qw_result = result
            update_progress(100, "Done!")
        else:
            st.error("Could not parse AI response. Please try again.")

    except Exception as e:
        st.error(f"Scan failed: {str(e)}")

    time.sleep(0.5)
    progress.empty()
    status.empty()

# ─── Results ───
if st.session_state.qw_result:
    data = st.session_state.qw_result
    wins = data.get("quick_wins", [])
    health = data.get("health_snapshot", {})
    scan = data.get("scan_summary", {})

    st.markdown('<div class="qw-divider"></div>', unsafe_allow_html=True)

    # Stats row
    total = health.get("total_issues", len(wins))
    crit = health.get("critical", 0)
    high = health.get("high", 0)
    med = health.get("medium", 0)
    low = health.get("low", 0)
    analyzed = scan.get("urls_analyzed", 0)

    st.markdown(f"""
    <div class="qw-stats">
        <div class="qw-stat">
            <div class="qw-stat-value">{analyzed}</div>
            <div class="qw-stat-label">Pages scanned</div>
        </div>
        <div class="qw-stat">
            <div class="qw-stat-value">{len(wins)}</div>
            <div class="qw-stat-label">Quick wins</div>
        </div>
        <div class="qw-stat qw-stat-critical">
            <div class="qw-stat-value">{crit}</div>
            <div class="qw-stat-label">Critical</div>
        </div>
        <div class="qw-stat qw-stat-high">
            <div class="qw-stat-value">{high}</div>
            <div class="qw-stat-label">High</div>
        </div>
        <div class="qw-stat qw-stat-medium">
            <div class="qw-stat-value">{med}</div>
            <div class="qw-stat-label">Medium</div>
        </div>
        <div class="qw-stat qw-stat-low">
            <div class="qw-stat-value">{low}</div>
            <div class="qw-stat-label">Low</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Tabs
    tab_wins, tab_fixes, tab_download = st.tabs(["⚡ Quick Wins", "🔧 Generate Fixes", "📥 Download"])

    with tab_wins:
        # Filter
        sev_filter = st.multiselect(
            "Filter by severity",
            ["critical", "high", "medium", "low"],
            default=["critical", "high", "medium", "low"],
        )
        filtered = [w for w in wins if (w.get("severity") or "medium").lower() in sev_filter]

        if not filtered:
            st.info("No quick wins match the selected filters.")
        else:
            for i, qw in enumerate(filtered):
                st.markdown(render_card(qw, i), unsafe_allow_html=True)

                example_urls = qw.get("example_urls", [])
                if example_urls:
                    with st.expander(f"Example URLs ({len(example_urls)})"):
                        for eu in example_urls[:5]:
                            st.code(eu, language=None)

    with tab_fixes:
        st.markdown("Select a quick win with auto-fix support to generate concrete fixes.")

        fixable = [w for w in wins if w.get("can_generate_fix")]

        if not fixable:
            st.info("No auto-fixable quick wins found in this scan.")
        else:
            options = {f"{w['title']} ({w['category']})": w for w in fixable}
            selected = st.selectbox("Choose a quick win to fix", list(options.keys()))

            if st.button("✨ Generate Fixes", use_container_width=True):
                win = options[selected]
                cat = win.get("category", "")

                # Build affected pages from crawl context
                crawl = st.session_state.crawl_context
                pages = crawl.get("pages", []) if crawl else []

                affected = []
                for pg in pages:
                    url = pg.get("final_url") or pg.get("url", "")
                    if cat == "titles":
                        val = pg.get("title", "")
                        if not val or val in [e.get("value") for e in (crawl or {}).get("examples", {}).get("duplicate_titles", [])]:
                            affected.append({"url": url, "current_value": val or "(missing)", "word_count": pg.get("word_count", 0)})
                    elif cat == "metas":
                        val = pg.get("meta", "")
                        if not val or val in [e.get("value") for e in (crawl or {}).get("examples", {}).get("duplicate_meta", [])]:
                            affected.append({"url": url, "current_value": val or "(missing)", "word_count": pg.get("word_count", 0)})
                    elif cat == "h1":
                        if pg.get("h1_count", 0) == 0:
                            affected.append({"url": url, "current_value": "(missing)", "title": pg.get("title", "")})
                    elif cat == "images":
                        if pg.get("images_missing_alt", 0) > 0:
                            affected.append({"url": url, "current_value": f"{pg.get('images_missing_alt', 0)} images missing alt", "title": pg.get("title", "")})

                if not affected:
                    affected = [{"url": eu, "current_value": "(see page)"} for eu in win.get("example_urls", [])[:10]]

                affected = affected[:15]  # cap

                with st.spinner("Generating fixes..."):
                    result = generate_fixes(cat, affected)

                if result and "fixes" in result:
                    st.session_state.fixes[cat] = result["fixes"]

                    for fix in result["fixes"]:
                        st.markdown(f"""
                        <div class="fix-item">
                            <div class="fix-url">{fix.get('url', '')}</div>
                            <div class="fix-current">✗ {fix.get('current_value', '')}</div>
                            <div class="fix-suggested">✓ {fix.get('suggested_fix', '')}</div>
                            <div class="fix-reason">{fix.get('reasoning', '')}</div>
                        </div>
                        """, unsafe_allow_html=True)
                else:
                    st.warning("Could not generate fixes. Try again.")

    with tab_download:
        st.markdown("Download your quick wins and generated fixes as an Excel file.")

        excel = create_excel(data, st.session_state.fixes if st.session_state.fixes else None)
        domain = data.get("domain", "site")
        fname = f"QuickWins_{domain}_{datetime.now().strftime('%Y%m%d')}.xlsx"

        st.download_button(
            label="📥 Download Excel Report",
            data=excel,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        st.caption("The Excel file includes the full quick wins list plus any fixes you generated.")


# Footer
st.markdown("""
<div class="qw-footer">
    Quick Wins · SEO Agent · Powered by Gemini
</div>
""", unsafe_allow_html=True)
